'''
Copyright 2022 Airbus SAS
Modifications on 2024/05/16-2024/07/10 Copyright 2024 Capgemini

Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

    http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
'''

import datetime

import openpyxl
import textdistance
from openpyxl.formatting.rule import FormulaRule, Rule
from openpyxl.worksheet.table import Table, TableStyleInfo


class SoSTerminology:
    """Class to use the SoS Terminology"""

    def __init__(self, xlFilePath, loadOrCreate='load'):
        """Constructor"""
        self.filePath = xlFilePath
        if loadOrCreate == 'load':
            self.workbook = openpyxl.load_workbook(xlFilePath)

    def get_sheet_headers(self, sheetName, columnsDict=None, maxColumn=0, stopAtNone=0):
        sheet = self.workbook[sheetName]
        if maxColumn > 0 and maxColumn < sheet.max_column:
            lastColumn = maxColumn
        else:
            lastColumn = sheet.max_column
        headers = [
            column[0].value
            for column in sheet.iter_cols(1, lastColumn)
            if stopAtNone == 0 or (stopAtNone == 1 and column[0].value is not None)
        ]

        if columnsDict is not None:
            for columnDict in columnsDict.values():
                if columnDict['header'] not in headers:
                    print(
                        f'{columnDict["header"]} is defined as a header but is not found in the Excel sheet',
                    )
            columnsList = [
                columnDict['header'] for columnDict in columnsDict.values()
            ]
            for header in headers:
                if header not in columnsList:
                    print(
                        f'{header} is found as a header in the Excel sheet but it is not defined in the code',
                    )
        return headers

    def get_sheet_dict(self, sheetName, headers, nameDictKey, startingRow=2):
        sheetDict = {}
        sheet = self.workbook[sheetName]
        indexDictKey = headers.index(nameDictKey) if nameDictKey != '' else 1

        for i, row in enumerate(
            sheet.iter_rows(min_row=startingRow, max_col=len(headers), values_only=True),
        ):
            isNone = True
            row_dict = {}
            NoneType = type(None)
            for j in range(len(headers)):
                if not isinstance(row[j], (str, NoneType)):
                    if isinstance(row[j], (datetime.datetime)):
                        row_dict[headers[j]] = row[j].strftime('%d %B %Y')
                else:
                    row_dict[headers[j]] = row[j]
                if row[j] is not None and row[j] != '' and str(row[j])[0:1] != '=':
                    isNone = False
            if not isNone:
                if nameDictKey != '':
                    sheetDict[row[indexDictKey]] = row_dict
                else:
                    sheetDict[indexDictKey] = row_dict
                    indexDictKey += 1

        return sheetDict

    def array_to_string(self, arrayToConvert):
        if arrayToConvert is not None:
            if isinstance(arrayToConvert, list):
                for v in arrayToConvert:
                    if v is None:
                        arrayToConvert.remove(v)
                    if v == 'null':
                        arrayToConvert.remove(v)
                if arrayToConvert is not None and len(arrayToConvert) > 0:
                    return ',\n'.join([str(i) for i in arrayToConvert])
                else:
                    return ''
            elif isinstance(arrayToConvert, (int, float, dict, str)):
                return str(arrayToConvert)
            else:
                print(f'Unknown type for {arrayToConvert}')
                return arrayToConvert
        else:
            return ''

    def get_most_similar_values(self, stringA, listString):
        similar = []
        stringA = stringA.replace('_dict', '').replace('_df', '')
        for stringB in listString:
            stringB = stringB.replace('_dict', '').replace('_df', '')
            if stringA != stringB:
                similarity1 = textdistance.damerau_levenshtein.similarity(
                    stringA.lower(), stringB.lower(),
                )
                similarity2 = textdistance.lcsseq.normalized_similarity(
                    stringA.lower(), stringB.lower(),
                )
                if similarity1 < 11 and similarity2 > 0.7:
                    similar.append(stringB)
        return similar

    def get_aggregate_param(self, paramA, paramDict):
        if paramA.find('_dict') > -1 or paramA.find('_df') > -1:
            paramSource = paramA.replace('_dict', '').replace('_df', '')
            if paramSource in paramDict:
                return paramSource
            else:
                return ''

    def create_sheet(self, title, sheetIndex):
        if title in self.workbook.sheetnames:
            old_ws_title = title + ' Backup'
            if old_ws_title in self.workbook.sheetnames:
                self.workbook.remove(self.workbook[old_ws_title])
            self.workbook[title].title = old_ws_title
            backupSheet = self.workbook[old_ws_title]
            index = self.workbook._sheets.index(backupSheet)
            nbSheet = len(self.workbook.sheetnames)
            self.workbook.move_sheet(self.workbook[old_ws_title], nbSheet - index)
            self.workbook[old_ws_title].sheet_state = 'hidden'
        return self.workbook.create_sheet(title=title, index=sheetIndex)

    def write_headers(self, sheet, headers):
        for col, header in enumerate(headers):
            sheet.cell(column=col + 1, row=1, value=header)

    def write_dict_values(self, sheet, dictToWrite, startRow):
        for row, rowDict in enumerate(dictToWrite):
            for col, key in enumerate(dictToWrite[rowDict]):
                sheet.cell(
                    column=col + 1, row=row + startRow, value=dictToWrite[rowDict][key],
                )

    def add_xl_table(self, tableName, sheet):
        tableDict = self.get_all_tables()
        tableName = tableName.replace(' ', '')

        if tableName in tableDict:
            tableSheet = self.workbook[tableDict[tableName]['worksheet']]
            del tableSheet.tables[tableName]

        # Define Excel Table
        tab = Table(displayName=tableName, ref=sheet.dimensions)

        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(
            name="TableStyleLight8",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style
        sheet.add_table(tab)

    def get_all_tables(self):
        tables_dict = {}
        # Go through each worksheet in the workbook
        for ws_name in self.workbook.sheetnames:
            ws = self.workbook[ws_name]
            # Get each table in the worksheet
            for tbl in ws.tables.values():
                tables_dict[tbl.name] = {
                    'table_name': tbl.name,
                    'table_displayName': tbl.displayName,
                    'worksheet': ws_name,
                }

        return tables_dict

    def set_columns_width(self, sheet):
        dims = {}
        for row in sheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max(
                        (dims.get(cell.column_letter, 0), len(str(cell.value))),
                    )
        for col, value in dims.items():
            maxWidth = 50
            if value > maxWidth:
                sheet.column_dimensions[col].width = maxWidth
            else:
                sheet.column_dimensions[col].width = value

    def add_formatting(self, sheet, column, formatType, formatting='', formula=''):
        start = openpyxl.utils.cell.absolute_coordinate(
            sheet.cell(2, column).coordinate,
        )
        end = openpyxl.utils.cell.absolute_coordinate(
            sheet.cell(sheet.max_row, column).coordinate,
        )
        rangeString = start + ':' + end

        if formatType == 'formula':
            rule = FormulaRule(formula=[formula], stopIfTrue=True, fill=formatting)
        elif formatType == 'containsBlanks':
            rule = Rule(
                type=formatType,
                formula=['LEN(TRIM(' + sheet.cell(2, column).coordinate + '))=0'],
                dxf=formatting,
            )
        elif formatType == 'notContainsBlanks':
            rule = Rule(
                type=formatType,
                formula=['LEN(TRIM(' + sheet.cell(2, column).coordinate + '))>0'],
                dxf=formatting,
            )
        else:
            raise Exception(f"Unhandled format type {formatType}")

        # Value must be one of
        # {'containsText', 'endsWith', 'colorScale',
        # 'notContainsText', 'containsBlanks' = 'LEN(TRIM(B2))=0'', 'notContainsBlanks = ['LEN(TRIM(B2))>0'],
        # 'cellIs', 'iconSet', 'uniqueValues', 'expression', 'containsErrors',
        # 'timePeriod', 'notContainsErrors', 'aboveAverage', 'top10', 'duplicateValues',
        # 'dataBar', 'beginsWith'}

        sheet.conditional_formatting.add(rangeString, rule)

    def add_list_to_sheet(self, listToAdd, sheetName, columnName):
        if len(listToAdd) and sheetName in self.workbook.sheetnames:
            sheet = self.workbook[sheetName]
            headers = self.get_sheet_headers(sheetName)
            if columnName in headers:
                col_index = headers.index(columnName)
                # retrieve first empty row
                rowCount = 0
                for max_row, row in enumerate(sheet, 1):
                    if row[col_index].value is not None:
                        rowCount += 1
                rowCount += 1

                # add elements to the sheet
                for element in listToAdd:
                    cell = sheet.cell(
                        column=col_index + 1, row=rowCount, value=str(element),
                    )
                    rowCount += 1

    def update_list_in_sheet(self, listToUpdate, sheetName, columnName):
        if len(listToUpdate) and sheetName in self.workbook.sheetnames:
            sheet = self.workbook[sheetName]
            headers = self.get_sheet_headers(sheetName)
            if columnName in headers:
                col_index = headers.index(columnName)

                # clear list:
                rowCount = 0
                for max_row, row in enumerate(sheet, 1):
                    if rowCount > 0 and row[col_index].value is not None:
                        cell = sheet.cell(
                            column=col_index + 1, row=rowCount + 1, value=None,
                        )
                        rowCount += 1
                # write new list
                rowCount = 2
                for element in listToUpdate:
                    cell = sheet.cell(
                        column=col_index + 1, row=rowCount + 1, value=element,
                    )
                    rowCount += 1
